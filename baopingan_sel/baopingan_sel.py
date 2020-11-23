import time
from selenium import webdriver
import datetime
from openpyxl import load_workbook

def get_info():
    username = []
    password = []
    wb = load_workbook(filename='C:\\Users\\盛意以山河\\Desktop\\info.xlsx') #自行更改绝对路径
    sheetnames = wb.get_sheet_names() 
    sheet = wb.get_sheet_by_name(sheetnames[0]) 
    row = wb[sheetnames[0]].max_row
    for rowNum in range(2,row+1):
        username.append(sheet.cell(row=rowNum,column=1).value) 
        password.append(sheet.cell(row=rowNum,column=2).value)
    return username, password
    

def bao(username, password):
    finshed = 0 #初始化数据
    sum = len(username)
    print('账号密码读取成功，共',sum,'人')

    #打开浏览器
    driver = webdriver.Chrome() 
    driver.get('http://yiqing.ctgu.edu.cn/wx/health/main.do') 
    time.sleep(2) 
    for x in range (sum):       
        driver.find_element_by_id('username1').click() #点击用户名输入框
        driver.find_element_by_id('username1').clear() #清空输入框
        driver.find_element_by_id('username1').send_keys(username[x]) #自动敲入用户名
        driver.find_element_by_id('password1').click() #点击密码输入框
        driver.find_element_by_id('password1').clear() #清空输入框
        driver.find_element_by_id('password1').send_keys(password[x]) #自动敲入密码
        driver.find_element_by_xpath('/html/body/main/section[2]/form/div[3]/input').click() #点击登陆
        time.sleep(5)
        #判断是是否密码错误
        try:
            tryfind = driver.find_element_by_id('msg')
            print('用户'+str(username[x])+'密码错误，请检查该用户密码后重试!\n')
            continue               
        except:
            print('用户'+str(username[x])+'登陆成功')
                
        #班委不一样,先切换角色
        s = driver.find_elements_by_id('menu_wjdc')#为什么能找到我也不清楚
        if len(s) != 0:
            print('该用户是班委')
            driver.find_element_by_id('menu_my').click() #点击我的按钮
            time.sleep(1)                
            driver.find_element_by_id('change_btn').click()#点击切换角色
            time.sleep(1)
            driver.find_element_by_xpath('/html/body/main/div[2]/section/form/div/div/div/div/label[2]').click()#点击学生
            driver.find_element_by_xpath('/html/body/main/div[2]/div/button').click()#点击确定
            time.sleep(5)  
        else:
            pass
        driver.find_element_by_xpath('/html/body/main/section/header/div[2]/button').click() #点击报平安
        # 尝试保平安，因为可能已经报过了
        try:
            driver.find_element_by_xpath('/html/body/main/div[1]/button').click()#点击报平安
            time.sleep(2)
            driver.find_element_by_xpath('/html/body/div[2]/div[3]/a[2]').click() #点击提交
            time.sleep(3)
        except:
            pass
        state1 = driver.find_element_by_xpath('/html/body/main/section/header/div[1]/span').text #判断是否报成功
        state2 = driver.find_element_by_xpath('/html/body/main/section/header/div[1]').text #获取真实名字
        if state1 == '今日已上报':
            print('用户'+str(username[x])+state2+'\n')
            finshed += 1
        else:
            print('用户',str(username[x]),'报平安失败，请重新上报')
            continue
        driver.find_element_by_id('menu_my').click() #点击我的按钮
        time.sleep(1)
        driver.find_element_by_id('submit_btn').click() #点击退出登录
        time.sleep(2)
    print('\n*****应报'+str(sum)+'人'+'实报'+str(finshed)+'人'+'请查收！*****')
 
def main():
    nowtime = datetime.datetime.now().strftime('%m-%d')#获取当前时间
    print('今天是'+nowtime)
    username, password = get_info()
    bao(username, password)

if __name__ == "__main__":
    main()