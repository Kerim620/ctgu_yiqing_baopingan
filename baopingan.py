import requests
import time
from bs4 import BeautifulSoup
from openpyxl import load_workbook


# 从excel文件中读取数据
def get_info():
    username = []
    password = []
    wb = load_workbook(filename='info.xlsx')
    sheetnames = wb.get_sheet_names()
    sheet = wb.get_sheet_by_name(sheetnames[0])
    row = wb[sheetnames[0]].max_row
    for rowNum in range(2, row + 1):
        username.append(sheet.cell(row=rowNum, column=1).value)
        password.append(sheet.cell(row=rowNum, column=2).value)
    sum = len(username)
    wb.close()
    return username, password, sum


# 登录，为后续做准备
def login(s, username, password):
    login_url = 'http://yiqing.ctgu.edu.cn/wx/index/loginSubmit.do'
    data = {'username': username, 'password': password}
    r = s.post(url=login_url, data=data)
    r.encoding = 'utf-8'
    return r.text


# 获取要带有data的html
def get_student_info(s):
    student_info_url = 'http://yiqing.ctgu.edu.cn/wx/health/toApply.do'
    r = s.get(url=student_info_url)
    r.encoding = 'utf-8'
    return r.text


# 解析html,获得data数据
def student_info_parse(html):
    bs = BeautifulSoup(html, 'lxml')
    data = {
        'ttoken': bs.find(attrs={'name': 'ttoken'})['value'],  # ttoken
        'province': '湖北省',  # 省
        'city': '宜昌市',  # 市
        'district': '西陵区',  # 县
        'adcode': '443000',  # 邮编
        'longitude': bs.find(attrs={'name': 'longitude'})['value'],  # 经度
        'latitude': bs.find(attrs={'name': 'latitude'})['value'],  # 纬度
        'sfqz': bs.find(attrs={'name': 'sfqz'})['value'],
        'sfys': bs.find(attrs={'name': 'sfys'})['value'],
        'sfzy': bs.find(attrs={'name': 'sfzy'})['value'],
        'sfgl': bs.find(attrs={'name': 'sfgl'})['value'],
        'status': bs.find(attrs={'name': 'status'})['value'],
        'szdz': '湖北省 宜昌市 西陵区',
        'sjh': bs.find(attrs={'name': 'sjh'})['value'],
        'lxrxm': bs.find(attrs={'name': 'lxrxm'})['value'],
        'lxrsjh': bs.find(attrs={'name': 'lxrsjh'})['value'],
        'sffr': '否',
        'sffrAm': '否',
        'sffrNoon': '否',
        'sffrPm': '否',
        'sffy': '否',
        'sfgr': '否',
        'qzglsj': '',
        'qzgldd': '',
        'glyy': '',
        'mqzz': '',
        'sffx': '否',
        'qt': '',
    }
    return data


# 向服务器post数据
def sent_info(s, data):
    sent_info_url = 'http://yiqing.ctgu.edu.cn/wx/health/saveApply.do'
    s.post(url=sent_info_url, data=data)


def main():
    username, password, sum = get_info()
    print('账号密码读取成功，共', str(sum), '人')
    finished = 0
    try:
        for i in range(sum):
            s = requests.Session()
            state = login(s, username[i], password[i])
            if state == 'success':
                print('用户', username[i], '登录成功')
            else:
                print('用户', username[i], '密码错误，登录失败\n')
                continue
            html = get_student_info(s)
            try:
                data = student_info_parse(html)
                sent_info(s, data)
                print('用户', username[i], '上报成功\n')
            except:
                print('用户', username[i], '今日已上报\n')
            s.close()
            finished += 1
            time.sleep(0.1)  # 每报完一个等待时间
    finally:
        print('应报:' + str(sum) + '实报:' + str(finished))


if __name__ == "__main__":
    main()
